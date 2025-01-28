import base64
import csv
import io
import openpyxl

from odoo import fields, api, _, models
from odoo.exceptions import UserError



class OpeningBalanceWizard(models.TransientModel):
    _name = 'opening.balance.wizard'
    _description = 'Wizard for Opening Balance Entry'

    file_data = fields.Binary(string="Template File", required=True)
    file_name = fields.Char(string="File Name")
    counterpart_account_id = fields.Many2one(
        'account.account', string="Counterpart Account", required=True,
        help="Account to balance the opening balance entry."
    )
    accounting_date = fields.Date(string="Accounting Date")

    @api.model
    def create_opening_entry(self, data, counterpart_account_id):
        """Create a balanced journal entry including the counterpart."""
        journal = self.env['account.journal'].search([('type', '=', 'general')], limit=1)
        if not journal:
            raise UserError(_("No general journal found."))

        # Calculate total debit and credit from the data
        total_debit = sum(line['debit'] for line in data)
        total_credit = sum(line['credit'] for line in data)

        # Determine the counterpart line's debit/credit to balance the entry
        if total_debit > total_credit:
            counterpart_debit = 0.0
            counterpart_credit = total_debit - total_credit
        else:
            counterpart_debit = total_credit - total_debit
            counterpart_credit = 0.0

        # Prepare all lines including the counterpart
        lines = [(0, 0, {
            'account_id': line['account_id'],
            'debit': line['debit'],
            'credit': line['credit'],
            'name': line['account_name'],
        }) for line in data]

        lines.append((0, 0, {
            'account_id': counterpart_account_id.id,
            'debit': counterpart_debit,
            'credit': counterpart_credit,
            'name': 'Counterpart Adjustment',
        }))

        # Create the journal entry
        move = self.env['account.move'].create({
            'journal_id': journal.id,
            'date': self.accounting_date,
            'move_type': 'entry',
            'line_ids': lines,
        })
        print(move)

        return move

    def action_process_template(self):
        """Read the uploaded .xlsx file and create a balanced opening balance entry."""
        if not self.file_data:
            raise UserError(_("Please upload a file."))

        if not self.counterpart_account_id:
            raise UserError(_("Please specify a counterpart account."))

        # Decode and load the .xlsx file
        try:
            file_content = base64.b64decode(self.file_data)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            sheet = workbook.active  # Read the first sheet
        except Exception as e:
            raise UserError(_("Error reading the uploaded file: %s") % str(e))

        # Validate headers
        expected_headers = ['Type','Account Code', 'Account', 'Debit', 'Credit']
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if headers != expected_headers:
            raise UserError(_("The uploaded file must have the following headers: %s") % ", ".join(expected_headers))

        # Read data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            print(row)
            ac_type,account_code, account_name, debit, credit = row
            if ac_type is None and account_code:
                account = self.env['account.account'].search([('code', '=', account_code)], limit=1)
                if not account:
                    raise UserError(_("Account with code %s not found.") % account_code)

                try:
                    data.append({
                        'account_id': account.id,
                        'account_name': account_name,
                        'debit': float(debit) if debit else 0.0,
                        'credit': float(credit) if credit else 0.0,
                    })
                except ValueError:
                    raise UserError(_("Invalid debit or credit value in row: %s") % row)

        # Create the journal entry with the counterpart
        self.create_opening_entry(data, self.counterpart_account_id)
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }