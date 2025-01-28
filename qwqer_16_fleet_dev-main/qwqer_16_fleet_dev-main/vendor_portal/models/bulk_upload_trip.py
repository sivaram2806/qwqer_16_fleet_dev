from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from openpyxl import load_workbook
import io
import base64


class BulkUploadTrip(models.Model):
    _name = 'bulk.upload.trip'
    _description = 'Bulk Upload Trips'
    _order = 'id desc'

    name = fields.Char(string="Name")
    upload_file = fields.Binary(string='Upload File')
    filename = fields.Char()
    line_ids = fields.One2many('bulk.upload.trip.line', 'bulk_upload_id', string='Trip Lines')
    log_data = fields.Json('Log Data')
    state = fields.Selection([
        ("draft", "Draft"),
        ("complete", "Completed"),
        ('complete_with_fail', 'Completed With Failures'),
        ("cancel", "Cancelled"),
    ], default="draft", string="Status")
    company_id = fields.Many2one(comodel_name='res.company', string="Company", required=True, readonly=True,
                                 default=lambda self: self.env.company)

    @api.model
    def create(self, vals):
        company_id = self.env.company.id
        sequence = self.env['ir.sequence'].search(
            [('company_id', '=', company_id), ('code', '=', 'bulk.upload.trip')])
        if not sequence:
            sequence = self.env['ir.sequence'].search([('code', '=', 'bulk.upload.trip')], limit=1)
            new_sequence = sequence.sudo().copy()
            new_sequence.company_id = company_id
            code = new_sequence.next_by_id()

        else:
            code = self.env['ir.sequence'].next_by_code('bulk.upload.trip')
        vals['name'] = code
        return super(BulkUploadTrip, self).create(vals)

    def create_lines(self, serial_no, state, vehicle_no, msg):
        self.env["bulk.upload.trip.line"].create({
            'name': serial_no,
            'state': state,
            'vehicle_no': vehicle_no,
            'desc': msg,
            'bulk_upload_id': self.id
        })

    def check_trips(self, row, vehicle_map):
        vehicle_pricing_lines = None
        msg = self.validate_line(row)
        if not msg:
            import_vehicle_no = row[2]
            vehicle_map_id = vehicle_map.get(import_vehicle_no)
            if not vehicle_map_id:
                msg = f"invalid vehicle {row[2]}."
            else:
                vehicle_pricing_lines = self.env['vehicle.pricing.line'].sudo().browse(int(vehicle_map_id))
            if not vehicle_pricing_lines:
                msg = f"Vehicle No {row[2]} does not exist.\nPlease contact fleet administrator"
        if msg:
            serial_no = int(row[0]) if row[0] else ''
            self.create_lines(serial_no, 'fail', row[2], msg)
            res_value = {'error_message': msg}
        else:
            res_value = {'vehicle_pricing_lines': vehicle_pricing_lines}
        return res_value

    def validate_line(self, row):
        field_list = ["SL No", "Trip Date", "Vehicle No", "comment", "Start Time", "End Time", "Start Odo", "End Odo"]
        missing_fields = [field_name for index, field_name in enumerate(field_list) if not row[index] and index != 3]
        if missing_fields:
            return self.format_missing_fields_message(missing_fields, row)
        return None

    def format_missing_fields_message(self, missing_fields, row):
        fields_str = ", ".join(missing_fields)
        if len(missing_fields) == 1:
            return f"{fields_str} column is mandatory for {'Vehicle No:' + row[2] if row[2] else 'SL No:' + str(row[0])}"
        return f"{fields_str}.\nAbove columns are mandatory for {'Vehicle No:' + row[2] if row[2] else 'SL No:' + str(row[0])}"

    def time_to_float_time(self, time_obj):
        try:
            float_time = time_obj.hour + time_obj.minute / 60 + time_obj.second / 3600
        except Exception:
            raise ValidationError("Invalid Time")
        return float_time

    def import_file(self):
        """ function to import trip sheet from  xlsx file """
        for record in self:
            if record.upload_file:
                # Decode the binary field
                excel_data = base64.b64decode(record.upload_file)
                # Use openpyxl to read the Excel file
                excel_file = io.BytesIO(excel_data)
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                mx_row = sheet.max_row
                # Find the last row with values
                last_value_row = max(
                    (index + 1 for index, row in enumerate(sheet.iter_rows(max_row=mx_row, values_only=True)) if
                     any(row)),
                    default=0
                )
                failed_status = False
                if last_value_row <= 1:
                    msg = 'Nothing to imported'
                    self.create_lines(None, 'fail', None, msg)
                    failed_status = True
                else:
                    vehicle_map = {}
                    if "Hidden" in workbook.sheetnames:
                        hidden_sheet = workbook["Hidden"]
                        for row in hidden_sheet.iter_rows(min_row=1, values_only=True):
                            vehicle_id, vehicle_number = row
                            vehicle_map[vehicle_number] = vehicle_id
                        for index, row in enumerate(sheet.iter_rows(max_row=last_value_row, values_only=True)):
                            if index == 0:
                                continue  # Skip the first row
                            check_constrain = self.check_trips(row, vehicle_map)
                            if not check_constrain.get('error_message'):
                                try:
                                    with self.env.cr.savepoint():
                                        trip_date = row[1].date()
                                        start_time = self.time_to_float_time(row[4])
                                        end_time = self.time_to_float_time(row[5])
                                        vendor = self.env.user.partner_id
                                        company_id = self.env.company
                                        vehicle_pricing_lines = check_constrain.get('vehicle_pricing_lines')
                                        if start_time < end_time and row[6] < row[7]:
                                            vals = {
                                                'vehicle_pricing_line_id': vehicle_pricing_lines.id or False,
                                                'vehicle_pricing_id': vehicle_pricing_lines.vehicle_pricing_id.id or False,
                                                'driver_name': vehicle_pricing_lines.driver_name or False,
                                                'vendor_id': vendor.id or False,
                                                'start_time': start_time or False,
                                                'end_time': end_time or False,
                                                'start_km': row[6] or 0.0,
                                                'end_km': row[7] or 0.0,
                                                'region_id': vehicle_pricing_lines.customer_id.region_id.id or False,
                                                'company_id': company_id.id or False
                                            }

                                            daily_trip = self.env['batch.trip.uh'].sudo().create({
                                                'trip_date': trip_date,
                                                'region_id': vehicle_pricing_lines.customer_id.region_id.id or None,
                                                'customer_id': vehicle_pricing_lines.customer_id.id or None,
                                                'frequency': vehicle_pricing_lines.customer_id.frequency or None,
                                                'sales_person_id': vehicle_pricing_lines.customer_id.order_sales_person.id or None,
                                                'is_vendor_trip': True,
                                                'comments': row[3] or False,
                                                'company_id': company_id.id or False,
                                                'batch_trip_uh_line_ids': [(0, 0, vals)] or False,
                                            })
                                        else:
                                            if start_time >= end_time and row[6] >= row[7]:
                                                msg = "Given Time and KM wrong"
                                            elif start_time >= end_time:
                                                msg = "Given start and end time wrong"
                                            elif row[6] >= row[7]:
                                                msg = "wrong Odo data"
                                            self.create_lines(row[0] or None, 'fail', row[2] or None, msg)
                                            raise ValidationError(msg)
                                except ValidationError as e:
                                    self.create_lines(row[0] or None, 'fail', row[2] or None, str(e))
                                    failed_status = True
                                except UserError as e:
                                    self.create_lines(row[0] or None, 'fail', row[2] or None, str(e))
                                    failed_status = True
                                except Exception:
                                    msg = "somthing went wrong!"
                                    self.create_lines(row[0] or None, 'fail', row[2] or None, msg)
                                    failed_status = True
                                else:
                                    msg = 'successfully imported'
                                    self.create_lines(row[0] or None, 'success', row[2] or None, msg)
                    else:
                        msg = 'Uploaded file is invalid, please upload correct one'
                        self.create_lines(None, 'fail', None, msg)
                        raise ValidationError(msg)
                if failed_status:
                    self.state = "complete_with_fail"
                else:
                    self.state = 'complete'

    def action_cancel(self):
        for rec in self:
            rec.state = 'cancel'


class BulkUploadTripLine(models.Model):
    _name = 'bulk.upload.trip.line'
    _description = 'Bulk Upload Trip Line'

    bulk_upload_id = fields.Many2one('bulk.upload.trip')
    name = fields.Char(string="Sl No")
    vehicle_no = fields.Char(string="Vehicle No")
    state = fields.Selection([('success', 'Success'),
                              ('fail', 'Failed')], string="Status")
    desc = fields.Text(string="Description")
